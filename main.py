import math
import os
import csv
from keras.models import Model
from keras.layers import Input
from keras.layers import Dense
from keras.layers import LSTM
from keras.layers import Dropout
from keras import Sequential
from keras.layers import BatchNormalization
import keras
import tensorflow as tf
import numpy as np
import matplotlib.pyplot as plt
from keras.utils.vis_utils import plot_model
import time
from tensorflow.python.client import device_lib
from tensorflow.python.framework import ops
from keras import backend as K
import openpyxl
from datetime import date
from operator import itemgetter
import keras_tuner as kt
# from kt import RandomSearch, Hyperband, BayesianOptimization
import random


def Deminov_model():
    '''
    тестирование модели Деминова
    '''

    path_data = "data/deminov/train675.csv"
    path_data_viborka_2 = "data/deminov/train870.csv"
    path_data_viborka_3 = "data/deminov/train102.csv"
    path_test_data = "data/deminov/input_parametrs.csv"
    train_input_data, train_output_data, input_data_test = get_data_deminov(path_data, path_test_data)
    norm_train, srednee = norm_deminov_srednee_otklon(train_input_data)
    mode = input("mode?\nt - train\np - predict\n--> ")

    if mode == "t":
        mode_train = input("use tuner?\ny - yes\nn - no\n--> ")
        if mode_train == "y":

            device = input("device?\nc - cpu\ng - gpu\n--> ")
            if device == "c":
                device = "/cpu:0"
            elif device == "g":
                device = "/gpu:0"
            else:
                print("select device")
                exit()

            # train_input_data, train_output_data, input_data_test = get_data_deminov(path_data, path_test_data)
            # norm_train, srednee = norm_deminov_srednee_otklon(train_input_data)
            ready_data = [np.array(norm_train), np.array(train_output_data) / 13.85]

            start = time.time()
            with tf.device(device):
                tuner = kt.RandomSearch(
                    build_model,  # функция создания модели
                    objective='val_accuracy',  # метрика, которую нужно оптимизировать -
                    # доля правильных ответов на проверочном наборе данных
                    max_trials=100,  # максимальное количество запусков обучения
                    directory='test_directory2'  # каталог, куда сохраняются обученные сети
                )
                tuner.search_space_summary()
                tuner.search(ready_data[0],  # Данные для обучения
                             ready_data[1],  # Правильные ответы
                             validation_split=0.2,
                             batch_size=128,
                             shuffle=True,
                             epochs=300)  # Количество эпох обучения
                tuner.results_summary()
                models = tuner.get_best_models(num_models=3)
            print(f'{device} time took: {time.time() - start:.4f}')
        elif mode_train == "n":

            device = input("device?\nc - cpu\ng - gpu\n--> ")
            if device == "c":
                device = "/cpu:0"
            elif device == "g":
                device = "/gpu:0"
            else:
                print("select device")
                exit()

            train_input_data, train_output_data, input_data_test = get_data_deminov(path_data, path_test_data)
            norm_train, srednee = norm_deminov_srednee_otklon(train_input_data)
            ready_data = [np.array(norm_train), np.array(train_output_data) / 13.85]
            start = time.time()

            with tf.device(device):
                model = keras.Sequential([
                    Dense(41, input_dim=len(ready_data[0][0]), activation='softplus', name='dense_1_input'),
                    Dense(20, activation='softplus', name='dense_2'),
                    Dense(10, activation='softplus', name='dense_3'),
                    Dense(2, activation='softmax', name='dense_4'),
                    #Dense(1, activation='softmax', name='dense_4'),
                    Dense(1, activation='linear', name='predict')
                ])
                model.compile(loss='mse', optimizer=tf.keras.optimizers.SGD(learning_rate=0.01), metrics=["acc"])

            with tf.device(device):
                history = model.fit(ready_data[0], ready_data[1], epochs=3000, verbose=1, validation_split=0.1, shuffle=True)

            print(f'{device} time took: {time.time() - start:.4f}')

            model.save("modeles/deminov.pb")
            plot_model(model, to_file='modeles/deminov.pb/model_1.png')
            plt.plot(history.history['loss'], label='loss')
            plt.plot(history.history['val_loss'], label='val_loss')
            plt.plot(history.history['acc'], label='acc')
            plt.plot(history.history['val_acc'], label='val_acc')
            plt.legend()
            plt.grid(True)
            plt.savefig(f"modeles/deminov.pb/MSE_{date.today()}.png")
            plt.show()

    elif mode == "p":
        device = input("device?\nc - cpu\ng - gpu\n--> ")
        if device == "c":
            device = "/cpu:0"
        elif device == "g":
            device = "/gpu:0"
        else:
            print("select device")
            exit()

        model = keras.models.load_model("modeles/deminov.pb")
        test_prepared_data_input, test_prepared_data_output, asd = get_data_deminov(path_data_viborka_2, path_test_data)
        test_prepared_data_input1, test_prepared_data_output1, asd = get_data_deminov(path_data_viborka_3, path_test_data)
        train_prepared_data_input, train_prepared_data_output, asd = get_data_deminov(path_data, path_test_data)
        asd = np.array(asd)
        with tf.device(device):
            #

            # for inpt, outpt in zip(train_prepared_data_input, train_prepared_data_output):
            #     data_test.append([(model.predict([[((inpt[0] - srednee[0][0]) / srednee[0][1]),
            #                                        ((inpt[1] - srednee[1][0]) / srednee[1][1]),
            #                                        ((inpt[2] - srednee[2][0]) / srednee[2][1]),
            #                                        ((inpt[3] - srednee[3][0]) / srednee[3][1])]])[0][0] * 13.85),
            #                       outpt])
            #
            # x1, x2 = [], []
            #
            # for i in range(len(data_test)):
            #     x1.append(data_test[i][0])
            #     x2.append(data_test[i][1])
            #
            # plt.plot(x1, color='r', label='data_predict')
            # plt.plot(x2, color='g', label='data_wait')
            # plt.title("train_input")
            # plt.legend()
            # plt.savefig(f"modeles/deminov.pb/predict/train_input.png", dpi=300)
            # plt.show()
            #
            # x1, x2 = [], []
            # sorted(data_test)
            # for i in range(len(data_test)):
            #     x2.append(data_test[i][0])
            #     x1.append(data_test[i][1])
            #
            # plt.plot(x2, color='r', label='data_predict')
            # plt.plot(x1, color='g', label='data_wait')
            # plt.title("test_input_not_sorted")
            # plt.legend()
            # plt.savefig(f"modeles/deminov.pb/predict/train_input_sorted.png")
            # plt.show()
            #
            #
            # # for i in range(len(data_test)):
            # #     print(i, data_test[i])
            #
            # data_test = []
            # for inpt, outpt in zip(test_prepared_data_input, test_prepared_data_output):
            #     # print([((inpt[0] - srednee[0][0]) / srednee[0][1]),
            #     #                                    ((inpt[1] - srednee[1][0]) / srednee[1][1]),
            #     #                                    ((inpt[2] - srednee[2][0]) / srednee[2][1]),
            #     #                                    ((inpt[3] - srednee[3][0]) / srednee[3][1])], inpt, outpt,
            #     #                                     [((inpt1[0] - srednee[0][0]) / srednee[0][1]),
            #     #                                    ((inpt1[1] - srednee[1][0]) / srednee[1][1]),
            #     #                                    ((inpt1[2] - srednee[2][0]) / srednee[2][1]),
            #     #                                    ((inpt1[3] - srednee[3][0]) / srednee[3][1])], inpt1, outpt1)
            #     data_test.append([(model.predict([[((inpt[0] - srednee[0][0]) / srednee[0][1]),
            #                                        ((inpt[1] - srednee[1][0]) / srednee[1][1]),
            #                                        ((inpt[2] - srednee[2][0]) / srednee[2][1]),
            #                                        ((inpt[3] - srednee[3][0]) / srednee[3][1])]])[0][0]) * 13.85,
            #                       outpt])
            # #print(srednee)
            # x1, x2 = [], []
            # s_data_test = Sort(data_test)
            # for i in range(len(data_test)):
            #     x2.append(s_data_test[i][0])
            #     x1.append(s_data_test[i][1])
            #
            # plt.plot(x2, color='r', label='data_predict')
            # plt.plot(x1, color='g', label='data_wait')
            # plt.title("test_input_not_sorted")
            # plt.legend()
            # plt.savefig(f"modeles/deminov.pb/predict/test_input_not_sorted.png")
            # plt.show()
            #
            # x1, x2 = [], []
            #
            # for i in range(len(s_data_test)):
            #     x1.append(s_data_test[i][0])
            #     x2.append(s_data_test[i][1])
            #
            # plt.plot(x1, color='r', label='data_predict')
            # plt.plot(x2, color='g', label='data_wait')
            # plt.title("test_input_sorted")
            # plt.legend()
            # plt.savefig(f"modeles/deminov.pb/predict/test_input_sorted.png")
            # plt.show()

            for j in range(1, 9, 2):
                data_test = []
                for inpt in asd:
                    data_test.append([(model.predict([[((inpt[0] - srednee[0][0]) / srednee[0][1]),
                                                       ((inpt[1] - srednee[1][0]) / srednee[1][1]),
                                                       ((np.float64("0."+str(j)) - srednee[2][0]) / srednee[2][1]),
                                                       ((inpt[3] - srednee[3][0]) / srednee[3][1])]])[0][0]) * 13.85,
                                      inpt[3], np.float64("0."+str(j))])
                    # print([((inpt[0] - srednee[0][0]) / srednee[0][1]),
                    #                                    ((inpt[1] - srednee[1][0]) / srednee[1][1]),
                    #                                    ((np.float64("0."+str(j)) - srednee[2][0]) / srednee[2][1]),
                    #                                    ((inpt[3] - srednee[3][0]) / srednee[3][1])], inpt, j)

            #print(srednee)
                x1, x2 = [], []
                for i in range(len(data_test)):
                    #print(i, data_test[i])
                    x1.append(data_test[i][0])
                    x2.append(data_test[i][1])
                plt.plot(x1, x2, color=(float("0."+str(j)), float("0."+str(j)), float("0."+str(j))), label="data_predict_albedo0_"+str(j))
            plt.legend()
            plt.savefig(f"modeles/deminov.pb/predict/g_ga_abledo0_7.png")
            plt.show()


def norm_deminov_srednee_otklon(data_input):
    '''
    нормализация данных для модели Деминова
    по 2-му способу https://www.youtube.com/watch?v=rRDRlc7xolU&t=1s
    '''
    a = []
    b = []
    c = []
    d = []
    for elem in data_input:
        a.append(float(elem[0]))
        b.append(float(elem[1]))
        c.append(float(elem[2]))
        d.append(float(elem[3]))

    avg_a = sum(a) / len(a)
    avg_b = sum(b) / len(b)
    avg_c = sum(c) / len(c)
    avg_d = sum(d) / len(d)

    temp_a = 0
    for i in range(len(a)):
        temp_a += (a[i] - avg_a) ** 2

    s_a = math.sqrt((1 / (len(a) - 1)) * temp_a)

    temp_b = 0
    for i in range(len(b)):
        temp_b += (b[i] - avg_b) ** 2
    s_b = math.sqrt((1 / (len(b) - 1)) * temp_b)

    temp_c = 0
    for i in range(len(c)):
        temp_c += (c[i] - avg_c) ** 2
    s_c = math.sqrt((1 / (len(c) - 1)) * temp_c)

    temp_d = 0
    for i in range(len(d)):
        temp_d += (d[i] - avg_d) ** 2
    s_d = math.sqrt((1 / (len(d) - 1)) * temp_d)

    for i in range(len(data_input)):
        data_input[i][0] = (a[i] - avg_a) / s_a
        data_input[i][1] = (b[i] - avg_b) / s_b
        data_input[i][2] = (c[i] - avg_c) / s_c
        data_input[i][3] = (d[i] - avg_d) / s_d

    return data_input, [[avg_a, s_a], [avg_b, s_b], [avg_c, s_c], [avg_d, s_d]]


def norm_data_deminov_min_max(data_input, data_output):
    '''
    нормализация данных для модели Деминова
    по 1-му способу https://www.youtube.com/watch?v=rRDRlc7xolU&t=1s
    '''
    a = []
    b = []
    c = []
    d = []
    out = []
    for elem in data_input:
        a.append(float(elem[0]))
        b.append(float(elem[1]))
        c.append(float(elem[2]))
        d.append(float(elem[3]))

    for elem in data_output:
        out.append(float(elem[0]))

    min_a, max_a = min(a), max(a)
    min_b, max_b = min(b), max(b)
    min_c, max_c = min(c), max(c)
    min_d, max_d = min(d), max(d)
    min_output, max_output = min(out), max(out)

    min_max_input, min_max_output = [[min_a, max_a], [min_b, max_b], [min_c, max_c],
                                     [min_d, max_d]], [[min_output, max_output]]

    for i in range(len(data_input)):
        data_input[i][0] = ((a[i] - min_a) / (max_a - min_a))
        data_input[i][1] = ((b[i] - min_b) / (max_b - min_b))
        data_input[i][2] = ((c[i] - min_c) / (max_c - min_c))
        data_input[i][3] = ((d[i] - min_d) / (max_a - min_d))
        data_output[i][0] = ((out[i] - min_output) / (max_output - min_output))

    return min_max_input, min_max_output, data_input, data_output


def min_max_norm_data_deminov_test(input_data, output_data, min_max_input, min_max_output):
    a = []
    b = []
    c = []
    d = []
    out = []

    for i in min_max_input:
        print(i)

    for elem in input_data:
        a.append(float(elem[0]))
        b.append(float(elem[1]))
        c.append(float(elem[2]))
        d.append(float(elem[3]))

    for elem in output_data:
        out.append(float(elem[0]))

    for i in range(len(input_data)):
        input_data[i][0] = ((a[i] - min_max_input[0][0]) / (min_max_input[0][1] - min_max_input[0][0]))
        input_data[i][1] = ((b[i] - min_max_input[1][0]) / (min_max_input[1][1] - min_max_input[1][0]))
        input_data[i][2] = ((c[i] - min_max_input[2][0]) / (min_max_input[2][1] - min_max_input[2][0]))
        input_data[i][3] = ((d[i] - min_max_input[3][0]) / (min_max_input[3][1] - min_max_input[3][0]))
        output_data[i][0] = ((out[i] - min_max_output[0][0]) / (min_max_output[0][1] - min_max_output[0][0]))

    return input_data, output_data


def get_data_deminov(path, path2):
    '''
    парсинг данных для модели Деминова
    path = train102.csv, train675.csv, train870.csv
    path2 = input_parametrs.csv
    '''
    data_input = []
    data_test = []
    data_output = []
    temp = []
    with open(path, newline='') as csvfile:
        data = csv.reader(csvfile, delimiter=';', quotechar='|')
        for row in data:
            temp.append([float(row[0].replace(",", ".")),
                         float(str(row[1].replace(",", "."))),
                         float(row[2].replace(",", ".")),
                         float(row[3].replace(",", ".")),
                         float(row[4].replace(",", "."))])
    # random.shuffle(temp)

    for elem in temp:
        data_input.append([elem[1], elem[2], elem[3], elem[4]])
        data_output.append([elem[0]])

    with open(path2, newline='') as csvfile:
        data = csv.reader(csvfile, delimiter=';', quotechar='|')
        for row in data:
            data_test.append([float(str(row[0].replace(",", "."))),
                              float(row[1].replace(",", ".")),
                              float(row[2].replace(",", ".")),
                              float(row[3].replace(",", "."))])

    return data_input, data_output, data_test


def Sort(sub_li):
    '''
    сортировка по второму элементу для графиков
    '''
    l = len(sub_li)
    for i in range(0, l):
        for j in range(0, l - i - 1):
            if (sub_li[j][1] > sub_li[j + 1][1]):
                tempo = sub_li[j]
                sub_li[j] = sub_li[j + 1]
                sub_li[j + 1] = tempo
    return sub_li


def build_model(hp):
    '''
    функция создания модели для тюнера keras.tuner
    '''
    model = keras.Sequential()
    activation_choice = hp.Choice('activation', values=['elu', 'selu', 'tanh', 'relu', 'softplus', 'sigmoid', 'linear',
                                                        'softmax'])
    # model.add(BatchNormalization())
    model.add(Dense(units=hp.Int('units_input', min_value=2, max_value=1024, step=32),
                    activation=activation_choice, input_dim=4))
    for i in range(hp.Int("num_layers", 1, 10)):
        model.add(Dense(units=hp.Int('units_hidden_' + str(i), min_value=2, max_value=1024, step=32),
                        activation=activation_choice))
    model.add(Dense(1, activation='linear'))
    opt = hp.Choice('optimizer', values=['Adam', 'SGD'])
    model.compile(optimizer=opt, loss='mean_squared_error', metrics=['accuracy'])
    # model.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=0.01), loss='mean_squared_error',
    # metrics=['accuracy'])
    return model


def chunks(lst, n):
    '''
    функция разделения массива на n частей
    '''
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def get_big_data(path):
    '''
    парсинг данных для моей модели
    path = AOD_0075.xlsx
    '''
    train_data = []
    test_data = []
    all_data = []
    # ALL_BIG_DATA_RETURN = []
    wb_obj = openpyxl.load_workbook(path)
    for name in wb_obj.sheetnames:
        temp_data = []
        sheet_obj = wb_obj[name]
        for i in range(0, 72, 9):
            for j in range(2, 8):
                temp_data.append(sheet_obj.cell(row=2, column=j + i).value)

        colums_data = list(chunks(temp_data, 6))

        for i in range(len(colums_data)):
            colums_data[i] = [colums_data[i][0], colums_data[i][1], colums_data[i][4], colums_data[i][5]]

        temp_data.clear()
        temp = []
        for j in range(0, 315, 38):
            for i in range(0, 72, 9):
                # for g in range(2, 8):
                #     temp.append(sheet_obj.cell(row=2, column=g + i).value)
                for h in range(13, 50):
                    for k in range(1, 10):
                        temp_data.append(sheet_obj.cell(row=h + j, column=k + i).value)

        colums_and_rows_data = list(chunks(temp_data, 9))

        j = 0
        for i in range(len(colums_and_rows_data)):
            if i > 0 and i % 37 == 0:
                j += 1
            if j > 7:
                j = 0
            colums_and_rows_data[i] = colums_and_rows_data[i] + colums_data[j]

        last_not_none = float(colums_and_rows_data[0][0].replace("ГА= ", "").replace(",", "."))
        for i in range(len(colums_and_rows_data)):
            if colums_and_rows_data[i][0] != None:
                colums_and_rows_data[i][0] = float(colums_and_rows_data[i][0].replace("ГА= ", "").replace(",", "."))
                last_not_none = float(colums_and_rows_data[i][0])
            else:
                colums_and_rows_data[i][0] = last_not_none

        for i in range(len(colums_and_rows_data)):
            colums_and_rows_data[i] = [colums_and_rows_data[i][0], colums_and_rows_data[i][2],
                                       colums_and_rows_data[i][8],
                                       colums_and_rows_data[i][9], colums_and_rows_data[i][10],
                                       colums_and_rows_data[i][11],
                                       colums_and_rows_data[i][12]]

        colums_and_rows_data = list(chunks(colums_and_rows_data, 37))

        for i in range(len(colums_and_rows_data)):
            colums_and_rows_data[i] = colums_and_rows_data[i][5:]

        colums_and_rows_data = list(chunks(colums_and_rows_data, 8))

        for i in range(len(colums_and_rows_data)):
            # print(i, colums_and_rows_data[i])
            all_data.append(colums_and_rows_data[i])
            if i % 2 == 0:
                train_data.append(colums_and_rows_data[i])
            else:
                test_data.append(colums_and_rows_data[i])

        # random.shuffle(train_data)
        # random.shuffle(test_data)
        # random.shuffle(all_data)
    return train_data, test_data, all_data


def prepare_data(a):
    '''
    подготовка данных
    перевод из str в  float
    формирование обучающей выборки data_input - входные данные, data_output - выходные данные
    '''
    data_input = []
    data_output = []
    print(len(a))
    for i in range(len(a)):
        data_input.append([float(a[i][0]), float(a[i][1]), float(a[i][2]), float(a[i][3]), float(a[i][4]), float(a[i][5])])
        data_output.append([float(a[i][6])])
        # data_input.append([float(a[i][1]), float(a[i][2]), float(a[i][3]), float(a[i][4]), float(a[i][5]), float(a[i][6])])
        # data_output.append([float(a[i][0])])  # TRAIN_DATA
            # for k in range(len(a[i][j])):
            #     data_output.append([float(a[i][j][k][0])])
            #     data_input.append([float(a[i][j][k][1]), float(a[i][j][k][2]), float(a[i][j][k][3]),
            #                        float(a[i][j][k][4]), float(a[i][j][k][5]),
            #                        float(a[i][j][k][6])])
                # data_input.append([float(a[i][j][k][1]), float(a[i][j][k][2]),
                #                    float(a[i][j][k][5]),
                #                    float(a[i][j][k][6])])

    return data_input, data_output


def norm_data(data):
    '''
    нормализация данных для моей модели
    по 1-му способу minmax https://www.youtube.com/watch?v=rRDRlc7xolU&t=1s
    '''

    norm_sca = []
    norm_rwi = []
    norm_wave = []
    norm_aod = []
    norm_sza = []
    norm_albedo = []

    for elem in data:
        norm_sca.append(float(elem[0]))
        norm_rwi.append(float(elem[1]))
        norm_wave.append(float(elem[2]))
        norm_aod.append(float(elem[3]))
        norm_sza.append(float(elem[4]))
        norm_albedo.append(float(elem[5]))

    min_norm_sca, max_norm_sca = min(norm_sca), max(norm_sca)
    min_norm_rwi, max_norm_rwi = min(norm_rwi), max(norm_rwi)
    min_norm_wave, max_norm_wave = min(norm_wave), max(norm_wave)
    min_norm_aod, max_norm_aod = min(norm_aod), max(norm_aod)
    min_norm_sza, max_norm_sza = min(norm_sza), max(norm_sza)
    min_norm_albedo, max_norm_albedo = min(norm_albedo), max(norm_albedo)

    for i in range(len(norm_sca)):
        norm_sca[i] = ((norm_sca[i] - min_norm_sca) / (max_norm_sca - min_norm_sca))
        norm_rwi[i] = ((norm_rwi[i] - min_norm_rwi) / (max_norm_rwi - min_norm_rwi))
        norm_wave[i] = ((norm_wave[i] - min_norm_wave) / (max_norm_wave - min_norm_wave))
        norm_aod[i] = ((norm_aod[i] - min_norm_aod) / (max_norm_aod - min_norm_aod))
        norm_sza[i] = ((norm_sza[i] - min_norm_sza) / (max_norm_sza - min_norm_sza))
        norm_albedo[i] = ((norm_albedo[i] - min_norm_albedo) / (max_norm_albedo - min_norm_albedo))

    for i in range(len(data)):
        data[i][0] = norm_sca[i]
        data[i][1] = norm_rwi[i]
        data[i][2] = norm_wave[i]
        data[i][3] = norm_aod[i]
        data[i][4] = norm_sza[i]
        data[i][5] = norm_albedo[i]

    return data, [[min_norm_sca, max_norm_sca], [min_norm_rwi, max_norm_rwi], [min_norm_wave, max_norm_wave], [min_norm_aod, max_norm_aod], [min_norm_sza, max_norm_sza], [min_norm_albedo, max_norm_albedo]]


def norm_srednee_otklon(data_input):
    '''
    нормализация данных для модели Деминова
    по 2-му способу https://www.youtube.com/watch?v=rRDRlc7xolU&t=1s
    '''
    a = []
    b = []
    c = []
    d = []
    e = []
    f = []
    for elem in data_input:
        a.append(float(elem[0]))
        b.append(float(elem[1]))
        c.append(float(elem[2]))
        d.append(float(elem[3]))
        e.append(float(elem[4]))
        f.append(float(elem[5]))

    avg_a = sum(a) / len(a)
    avg_b = sum(b) / len(b)
    avg_c = sum(c) / len(c)
    avg_d = sum(d) / len(d)
    avg_e = sum(e) / len(e)
    avg_f = sum(f) / len(f)

    temp_a = 0
    for i in range(len(a)):
        temp_a += (a[i] - avg_a) ** 2

    s_a = math.sqrt((1 / (len(a) - 1)) * temp_a)

    temp_b = 0
    for i in range(len(b)):
        temp_b += (b[i] - avg_b) ** 2
    s_b = math.sqrt((1 / (len(b) - 1)) * temp_b)

    temp_c = 0
    for i in range(len(c)):
        temp_c += (c[i] - avg_c) ** 2
    s_c = math.sqrt((1 / (len(c) - 1)) * temp_c)

    temp_d = 0
    for i in range(len(d)):
        temp_d += (d[i] - avg_d) ** 2
    s_d = math.sqrt((1 / (len(d) - 1)) * temp_d)

    temp_e = 0
    for i in range(len(e)):
        temp_e += (e[i] - avg_e) ** 2
    s_e = math.sqrt((1 / (len(e) - 1)) * temp_e)

    temp_f = 0
    for i in range(len(f)):
        temp_f += (f[i] - avg_f) ** 2
    s_f = math.sqrt((1 / (len(f) - 1)) * temp_f)

    for i in range(len(data_input)):
        data_input[i][0] = (a[i] - avg_a) / s_a
        data_input[i][1] = (b[i] - avg_b) / s_b
        data_input[i][2] = (c[i] - avg_c) / s_c
        data_input[i][3] = (d[i] - avg_d) / s_d
        data_input[i][4] = (e[i] - avg_e) / s_e
        data_input[i][5] = (f[i] - avg_f) / s_f

    return data_input, [[avg_a, s_a], [avg_b, s_b], [avg_c, s_c], [avg_d, s_d], [avg_e, s_e], [avg_f, s_f]]


def least_squares_method(y_true, y_pred):
    res = (y_true - y_pred)
    return res



def my_model():
    '''
    моя модель
    '''
    device = input("device?\nc - cpu\ng - gpu\n--> ")
    if device == "c":
        device = "/cpu:0"
    elif device == "g":
        device = "/gpu:0"
    else:
        my_model()

    # path = "TRAIN_DATA.csv"
    path = "srednee.csv"
    # path = "minmax.csv"
    data_for_prepare = []

    with open(path, newline='') as csvfile:
        data = csv.reader(csvfile, delimiter=',', quotechar='|')
        for row in data:
            data_for_prepare.append(row)

    minmax = [[float(data_for_prepare[0][0]),
               float(data_for_prepare[0][1])],
              [float(data_for_prepare[0][2]),
               float(data_for_prepare[0][3])],
              [float(data_for_prepare[0][4]),
               float(data_for_prepare[0][5])],
              [float(data_for_prepare[0][6]),
               float(data_for_prepare[0][7])],
              [float(data_for_prepare[0][8]),
               float(data_for_prepare[0][9])],
              [float(data_for_prepare[0][10]),
               float(data_for_prepare[0][11])]]

    print(minmax)
    del data_for_prepare[0]

    # for i in range(len(data_for_prepare)):
    #     print(i, data_for_prepare[i])

    train_data_input, train_data_output = prepare_data(data_for_prepare)
    # np.set_printoptions(suppress=True)
    # x_train = np.array(train_data_input)
    # print(x_train[:5])
    # mean = x_train.mean(axis=0)
    # std = x_train.std(axis=0)
    # x_train -= mean
    # x_train /= std
    #
    # print(x_train[:5])

    # for i, j in zip(train_data_input, train_data_output):
    #     print(i, j)
    #
    # print(minmax)


    # start = time.time()
    # tuner = kt.RandomSearch(
    #     build_model,  # функция создания модели
    #     objective='val_accuracy',  # метрика, которую нужно оптимизировать -
    #     # доля правильных ответов на проверочном наборе данных
    #     max_trials=10,  # максимальное количество запусков обучения
    #     directory='test_directory2'  # каталог, куда сохраняются обученные сети
    # )
    #
    # tuner.search_space_summary()
    # tuner.search(prepared_data_train_input,  # Данные для обучения
    #              prepared_data_train_output,  # Правильные ответы
    #              batch_size=8,
    #              # validation_data=(test_prepared_data_input, test_prepared_data_output),
    #              validation_split=0.3,
    #              epochs=500,  # Количество эпох обучения
    #              )
    #
    # tuner.results_summary()
    #
    # models = tuner.get_best_models(num_models=3)
    # i = 0
    # for model in models:
    #     model.build()
    #     model.summary()
    #     model.evaluate(test_prepared_data_input, test_prepared_data_output)
    #     print()
    #     model.save(f"C:\\Users\\27les\\PycharmProjects\\DIPLOM_TENSORFLOW\\model{i}.pb")
    #     i+=1
    #
    # print('time took: {0:.4f}'.format(time.time() - start))



    #
    # start = time.time()
    # with tf.device(device):
    #     model = keras.Sequential([
    #         # BatchNormalization(),
    #         Dense(1024, input_shape=(6,), activation='relu'),
    #         Dense(128, activation='relu'),
    #         # Dense(128, activation='relu'),
    #         # Dense(65, activation='relu'),
    #         # Dense(16, activation='relu'),
    #         # Dense(8, activation='relu'),
    #         # Dense(4, activation='relu'),
    #         # Dense(2, activation='relu'),
    #         # Dense(21, activation='softplus'),
    #         # Dense(17, activation='softplus'),
    #         # Dense(5, activation='softplus'),
    #         # Dense(1, activation='softplus'),
    #         # #Dense(1, activation='linear'),
    #         Dense(1, activation='linear')
    #     ])
    #     model.compile(loss='mse', optimizer=tf.keras.optimizers.Adam(0.01), metrics=['mae'])
    #
    #     # if path == "srednee.csv":
    #     #     print(model.predict([[((12.98 - minmax[0][0]) / minmax[0][1]),
    #     #                           ((0.286714 - minmax[1][0]) / minmax[1][1]),
    #     #                           ((0.675 - minmax[2][0]) / minmax[2][1]),
    #     #                           ((0.4 - minmax[3][0]) / minmax[3][1]),
    #     #                           ((60 - minmax[4][0]) / minmax[4][1]),
    #     #                           ((0.0 - minmax[5][0]) / minmax[5][1])]]))
    #     # elif path == "minmax.csv":
    #     #     print(model.predict([[((12.98 - minmax[0][0]) / (minmax[0][1] - minmax[0][0])),
    #     #                           ((0.286714 - minmax[1][0]) / (minmax[1][1] - minmax[1][0])),
    #     #                           ((0.675 - minmax[2][0]) / (minmax[2][1] - minmax[2][0])),
    #     #                           ((0.4 - minmax[3][0]) / (minmax[3][1] - minmax[3][0])),
    #     #                           ((60 - minmax[4][0]) / (minmax[4][1] - minmax[4][0])),
    #     #                           ((0.0 - minmax[5][0]) / (minmax[5][1] - minmax[5][0]))]]) )
    #
    # with tf.device(device):
    #     history = model.fit(x_train, np.array(train_data_output), epochs=2000, verbose=1, validation_split=0.1, batch_size=2048, shuffle=True)
    #
    # loss = model.evaluate(x_train, np.array(train_data_output), verbose=1)
    #
    # print(f'{device} time took: {time.time() - start:.4f}')
    # # if path == "srednee.csv":
    # #     print(model.predict([[((12.98 - minmax[0][0]) / minmax[0][1]),
    # #                         ((0.286714 - minmax[1][0]) / minmax[1][1]),
    # #                         ((0.675 - minmax[2][0]) / minmax[2][1]),
    # #                         ((0.4 - minmax[3][0]) / minmax[3][1]),
    # #                         ((60 - minmax[4][0]) / minmax[4][1]),
    # #                         ((0.0 - minmax[5][0]) / minmax[5][1])]]) * 13.85)
    # # elif path == "minmax.csv":
    # #     print(model.predict([[((12.98 - minmax[0][0]) / (minmax[0][1] - minmax[0][0])),
    # #                           ((0.286714 - minmax[1][0]) / (minmax[1][1] - minmax[1][0])),
    # #                           ((0.675 - minmax[2][0]) / (minmax[2][1] - minmax[2][0])),
    # #                           ((0.4 - minmax[3][0]) / (minmax[3][1] - minmax[3][0])),
    # #                           ((60 - minmax[4][0]) / (minmax[4][1] - minmax[4][0])),
    # #                           ((0.0 - minmax[5][0]) / (minmax[5][1] - minmax[5][0]))]]) * 13.85)
    # # else:
    # #     pass
    #
    # # #
    # model.save("modeles/my_model.pb")
    # plt.plot(history.history['loss'], label="loss")
    # plt.plot(history.history['mae'], label="mae")
    # plt.plot(history.history['val_loss'], label="val_loss")
    # plt.plot(history.history['val_mae'], label="val_mae")
    # plt.grid(True)
    # plt.legend()
    # plt.savefig(f"modeles/MSE_{date.today()}.png")
    # plt.show()



    model = keras.models.load_model("modeles/my_model.pb")
    # # loss = model.evaluate(np.array(train_data_input), np.array(train_data_output) * 13.85, verbose=1)


    for i in range(int(input("count: "))):
        print(model.predict([[((float(input("SCA: ")) - minmax[0][0]) / minmax[0][1]),
                              ((float(input("RWI: ")) - minmax[1][0]) / minmax[1][1]),
                              ((float(input("L: ")) - minmax[2][0]) / minmax[2][1]),
                              ((float(input("AOD: ")) - minmax[3][0]) / minmax[3][1]),
                              ((float(input("SZA: ")) - minmax[4][0]) / minmax[4][1]),
                              ((float(input("Albedo: ")) - minmax[5][0]) / minmax[5][1])]]))




    # if path == "srednee.csv":
    #     print(model.predict([[((12.98 - minmax[0][0]) / minmax[0][1]),
    #                         ((0.286714 - minmax[1][0]) / minmax[1][1]),
    #                         ((0.675 - minmax[2][0]) / minmax[2][1]),
    #                         ((0.4 - minmax[3][0]) / minmax[3][1]),
    #                         ((60 - minmax[4][0]) / minmax[4][1]),
    #                         ((0.0 - minmax[5][0]) / minmax[5][1])]]) )
    # elif path == "minmax.csv":
    #     print(model.predict([[((12.98 - minmax[0][0]) / (minmax[0][1] - minmax[0][0])),
    #                           ((0.286714 - minmax[1][0]) / (minmax[1][1] - minmax[1][0])),
    #                           ((0.675 - minmax[2][0]) / (minmax[2][1] - minmax[2][0])),
    #                           ((0.4 - minmax[3][0]) / (minmax[3][1] - minmax[3][0])),
    #                           ((60 - minmax[4][0]) / (minmax[4][1] - minmax[4][0])),
    #                           ((0.0 - minmax[5][0]) / (minmax[5][1] - minmax[5][0]))]]))
    # else:
    #     pass

    # for i in range(0, 665280, 665):
    #     if path == "srednee.csv":
    #         a = model.predict([[train_data_input[i][0], train_data_input[i][1], train_data_input[i][2], train_data_input[i][3], train_data_input[i][4], train_data_input[i][5]]])
    #         # a = model.predict([[((train_data_input[i][0] - minmax[0][0]) / minmax[0][1]),
    #         #                     ((train_data_input[i][1] - minmax[1][0]) / minmax[1][1]),
    #         #                     ((train_data_input[i][2] - minmax[2][0]) / minmax[2][1]),
    #         #                     ((train_data_input[i][3] - minmax[3][0]) / minmax[3][1]),
    #         #                     ((train_data_input[i][4] - minmax[4][0]) / minmax[4][1]),
    #         #                     ((train_data_input[i][5] - minmax[5][0]) / minmax[5][1])]])
    #         print(a, train_data_output[i][0] * 13.85)
    #     #elif path == "minmax.csv":
    #         # a = model.predict([[((train_data_input[i][0] - minmax[0][0]) / (minmax[0][1] - minmax[0][0])),
    #         #                     ((train_data_input[i][1] - minmax[1][0]) / (minmax[1][1] - minmax[1][0])),
    #         #                     ((train_data_input[i][2] - minmax[2][0]) / (minmax[2][1] - minmax[2][0])),
    #         #                     ((train_data_input[i][3] - minmax[3][0]) / (minmax[3][1] - minmax[3][0])),
    #         #                     ((train_data_input[i][4] - minmax[4][0]) / (minmax[4][1] - minmax[4][0])),
    #         #                     ((train_data_input[i][5] - minmax[5][0]) / (minmax[5][1] - minmax[5][0]))]])
    #         #print(a, a * 13.85, train_data_output[i][0] * 13.85)
    #     else:
    #         pass


def main():
    '''
    main func
    '''
    model = input("which model?\ndeminov - d\nmy model - m\n--> ")
    if model == "d":
        Deminov_model()
    elif model == "m":
        my_model()
    else:
        main()


if __name__ == "__main__":
    main()
