import os
import openpyxl
import time
import csv
path = "C:\\Users\\27les\\PycharmProjects\\DIPLOM_TENSORFLOW\\data\\PREPARE_DATA\\"
lstdir = os.listdir(path)


def chunks(lst, n):
    '''
    функция разделения массива на n частей
    '''
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def get_big_data(path):
    '''
    парсинг данных для моей модели
    path = AOD_0075.xlsx
    '''
    train_data = []
    test_data = []
    all_data = []
    # ALL_BIG_DATA_RETURN = []
    wb_obj = openpyxl.load_workbook(path)
    for name in wb_obj.sheetnames:
        temp_data = []
        if "Z" in name and "One_Z0" not in name and "templateZ" not in name:
            if "(2)" not in name:
                sheet_obj = wb_obj[name]
                for i in range(0, 135, 9):
                    for j in range(2, 8):
                        temp_data.append(sheet_obj.cell(row=2, column=j + i).value)
            else:
                sheet_obj = wb_obj[name]
                for i in range(0, 135, 9):
                    for j in range(2, 8):
                        temp_data.append(sheet_obj.cell(row=2, column=j + i).value)

            # for i in range(len(temp_data)):
            #     print(i, temp_data[i])

            colums_data = list(chunks(temp_data, 6))
            #
            for i in range(len(colums_data)):
                colums_data[i] = [colums_data[i][0], colums_data[i][1], colums_data[i][4], colums_data[i][5]]

            # for i in range(len(colums_data)):
            #     print(i, colums_data[i])
            temp_data.clear()
            temp = []
            for j in range(0, 315, 38):
                for i in range(0, 135, 9):
                    # for g in range(2, 8):
                    #     temp.append(sheet_obj.cell(row=2, column=g + i).value)
                    for h in range(13, 50):
                        for k in range(1, 10):
                            temp_data.append(sheet_obj.cell(row=h + j, column=k + i).value)

            colums_and_rows_data = list(chunks(temp_data, 9))

            j = 0
            for i in range(len(colums_and_rows_data)):
                if i > 0 and i % 37 == 0:
                    j += 1
                if j > 14:
                    j = 0
                colums_and_rows_data[i] = colums_and_rows_data[i] + colums_data[j]

            last_not_none = float(colums_and_rows_data[0][0].replace("ГА= ", "").replace(",", "."))
            for i in range(len(colums_and_rows_data)):
                if colums_and_rows_data[i][0] != None:
                    colums_and_rows_data[i][0] = float(colums_and_rows_data[i][0].replace("ГА= ", "").replace(",", "."))
                    last_not_none = float(colums_and_rows_data[i][0])
                else:
                    colums_and_rows_data[i][0] = last_not_none

            for i in range(len(colums_and_rows_data)):
                colums_and_rows_data[i] = [colums_and_rows_data[i][0], colums_and_rows_data[i][2],
                                           colums_and_rows_data[i][8],
                                           colums_and_rows_data[i][9], colums_and_rows_data[i][10],
                                           colums_and_rows_data[i][11],
                                           colums_and_rows_data[i][12]]

            colums_and_rows_data = list(chunks(colums_and_rows_data, 37))

            for i in range(len(colums_and_rows_data)):
                colums_and_rows_data[i] = colums_and_rows_data[i][5:]

            colums_and_rows_data = list(chunks(colums_and_rows_data, 8))


            # for i in range(len(colums_and_rows_data)):
            #     print(i, colums_and_rows_data[i])
            for i in range(len(colums_and_rows_data)):
                # # print(i, colums_and_rows_data[i])
                # all_data.append(colums_and_rows_data[i])
                # if i % 2 == 0:
                #     train_data.append(colums_and_rows_data[i])
                # else:
                test_data.append(colums_and_rows_data[i])

    return test_data #train_data, test_data, all_data


start = time.time()
asd = []
for i in lstdir:
    temp_lst_dir = os.listdir(path + i)
    for j in temp_lst_dir:
        print(path + i + "/" + j)
        #get_big_data(path + i + "/" + j)
        asd.append(get_big_data(path + i + "/" + j))  # One_Z0 templateZ

# u = 0
#
# for i in asd:
#     for j in i:
#         for h in j:
#             for k in range(len(h)):
#                 print(h[k])
with open("TRAIN_DATA.csv", "w", encoding='UTF8', newline='') as f:
    www = csv.writer(f)
    for i in asd:
        for j in i:
            for h in j:
                for k in range(len(h)):
                    www.writerow([str(h[k][0]), str(h[k][1]), str(h[k][2]), str(h[k][3]), str(h[k][4]), str(h[k][5]), str(h[k][6])])
                    #print(k, h[k], u)
                    #u += 1

print(time.time() - start)